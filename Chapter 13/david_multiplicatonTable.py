# Write your code here :-)
import sys
import openpyxl
from openpyxl.styles import Font

def create_multiplication_table(N):
    wb = openpyxl.Workbook()
    sheet = wb.active

    for i in range(1, N + 1):
        sheet.cell(row=1, column=i + 1, value=i).font = Font(bold=True)
        sheet.cell(row=i + 1, column=1, value=i).font = Font(bold=True)

    for row in range(2, N + 2):
        for col in range(2, N + 2):
            sheet.cell(row=row, column=col, value=(row - 1) * (col - 1))

    wb.save(f'C://Users//david//multiplication_table {N}.xlsx')
    print(f"Multiplication table for {N} saved as 'multiplication_table {N}.xlsx'.")

def main():
    if len(sys.argv) != 2:
        print("Usage: python multiplicationTable.py <N>")
        sys.exit(1)

    try:
        N = int(sys.argv[1])
        if N <= 0:
            print("Please provide a positive integer for N.")
            sys.exit(1)
        create_multiplication_table(N)
    except ValueError:
        print("Please provide a valid integer for N.")
        sys.exit(1)

main()
