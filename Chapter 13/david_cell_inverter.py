# Write your code here :-)
import openpyxl

wb = openpyxl.load_workbook('C://Users//david//documents//Inveter.xlsx')
sheet = wb.active

sheetData = []

for row in sheet.iter_rows(values_only=True):
    row_data = []
    for cell in row:
        row_data.append(cell)
    sheetData.append(row_data)

#print(sheetData)
sheetData1 = list(zip(*sheetData))
#print(sheetData1)

wb_new = openpyxl.Workbook()
newSheet = wb_new.active

for i, row in enumerate(sheetData1, start=1):
    #print(i)
    for j, value in enumerate(row, start=1):
        #print(j)
        newSheet.cell(row=i, column=j, value=value)

wb_new.save('C://Users//david//documents//Inveter1.xlsx')
print('DONE')
